from asyncio.windows_events import NULL
import requests
from openpyxl import Workbook
from datetime import date
import os
from pymongo import MongoClient
import shutil
from openpyxl import load_workbook

# date kısmı klasörlemede yardımcı olucak
now = date.today()
CURRENT_DATE = now.strftime("%d-%m-%Y")

# APi bilgileri
API_KEY = 'YOUR API KEYS'
CHANNEL_ID = "CHANNEL ID"
url = f'https://www.googleapis.com/youtube/v3/channels?part=snippet,statistics&id={CHANNEL_ID}&key={API_KEY}'

response = requests.get(url)

# tanımlamalar
titlename = "KANAL ADI"
goruntulenme = "viewCount"
videosayisi = "videoCount"
acıklama = "KANAL AÇIKLAMASI"
firstdatettanım = "BAŞLANGIÇ TARİHİ"
dizin = os.getcwd()

class YoutubeParse():
    def __init__(self):
        if os.path.exists(f"{dizin}\\{CURRENT_DATE}"):  # normal validate
            for (key, value) in response.json().items():
                if key == "items":
                    self.channelname = value[0]["snippet"]["title"]
                    self.about = value[0]["snippet"]["description"]
                    self.firstdate = value[0]["snippet"]["publishedAt"]
                    self.istatistik = value[0]["statistics"]
                    self.viewcount = self.istatistik[goruntulenme]
                    self.videocount = self.istatistik[videosayisi]
        else:
            os.makedirs(f"{dizin}\\{CURRENT_DATE}")
            self.validate()

    def validate(self):
            wb = Workbook()
            writerows = wb.active
            for row in range(1, 2):
                writerows["A" + str(row)] = titlename
                writerows["B" + str(row)] = goruntulenme
                writerows["C" + str(row)] = videosayisi
                writerows["D" + str(row)] = acıklama
                writerows["E" + str(row)] = firstdatettanım
                writerows["A" + str(row+1)] = self.channelname
                writerows["B" + str(row+1)] = self.viewcount
                writerows["C" + str(row+1)] = self.videocount
                writerows["D" + str(row+1)] = self.about
                writerows["E" + str(row+1)] = self.firstdate
            src = f"{CURRENT_DATE}\\{self.channelname}.xlsx"
            wb.save(src)

            # diğer tarafa yedekleme işlemi
            os.chdir(".")
            mydizin = f"youtube_yedek\\{CURRENT_DATE}"
            if os.path.exists(f"{mydizin}"):
                shutil.copy(src, f"{mydizin}\\{self.channelname}.xlsx")
            
            else:
                os.makedirs(f"{mydizin}")
                shutil.copy(src, f"{mydizin}\\{self.channelname}.xlsx")
            # yedekleme sonu

    def insertToDatabase(self):
        self.channelname = self.channelname
        myclient = MongoClient(
        "mongodb+srv://YOURDBINFO@cluster0.hclkh0g.mongodb.net/test")
        mydb = myclient["MyDatabase"]
        mycol = mydb["channels"]

        wb =load_workbook(f"{dizin}\\{CURRENT_DATE}\\{self.channelname}.xlsx")
        ws = wb.active

        titlename = ws.cell(row=1, column=1).value
        channelname = ws.cell(row=2, column=1).value
        goruntulenme = ws.cell(row=1, column=2).value
        viewcount = ws.cell(row=2, column=2).value 
        videosayisi = ws.cell(row=1, column=3).value
        videocount = ws.cell(row=2, column=3).value
        acıklama = ws.cell(row=1, column=4).value
        about = ws.cell(row=2, column=4).value
        firstdatettanım = ws.cell(row=1, column=5).value
        firstdate = ws.cell(row=2, column=5).value 

        myarray = [{"_id" : CHANNEL_ID, titlename: channelname, goruntulenme: viewcount, 
        videosayisi: videocount, acıklama: about, firstdatettanım: firstdate}]

        mycol.insert_many(myarray)
        ("okayy")

youtubecounter = YoutubeParse()
youtubecounter.validate()
youtubecounter.insertToDatabase()